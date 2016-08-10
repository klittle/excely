#!/usr/bin/env python

import sys, os

# read
from openpyxl import load_workbook
# write
from openpyxl import Workbook

from openpyxl.compat import range
# from openpyxl.cell import get_column_letter


def spelled_word(in_word):
    """
    returns spelled_word
    """
    return in_word + 'foo'


def read_in_write_out(in_filename, out_filename):
    """
    :param in_filename:
    excel file with extension .xlsx
    file to read from

    :param out_filename:
    excel file with extension .xlsx
    file to write to
    """
    in_workbook = load_workbook(in_filename)
    # ['Sheet 1']
    print(in_workbook.sheetnames)
    in_sheet = in_workbook.active
    print('in_sheet name ', in_workbook.sheetnames[0])

    out_workbook = Workbook()
    out_sheet = out_workbook.active
    out_sheet.title = "my_sheet"

    # http://stackoverflow.com/questions/37440855/how-do-i-iterate-through-cells-in-a-specific-column-using-openpyxl-1-6
    first_non_header_row = 2
    # column to read in in_sheet
    in_column = 2
    # column to write in out_sheet
    out_column = 1

    for row in range(first_non_header_row, in_sheet.max_row + 1):
        in_word = in_sheet.cell(row=row, column=in_column).value
        out_word = spelled_word(in_word)
        print(in_word + ', ' + out_word)
        # out_sheet set cell value
        _ = out_sheet.cell(column=out_column, row=row, value=out_word)

    out_workbook.save(filename = out_filename)


def write_spelling_matches_to_file(spellings_file_name, misspelled_file_name):
    """
    Reads from 2 excel files. Writes to misspelled_file_name

    :param spellings_file_name:
    excel file with extension .xlsx
    file to read from
    spellings number of rows is <= misspelled number of rows
    column b contains misspelled word.
    column b is sorted alphabetically ascending
    column d has correct spelling

    :param misspelled_file_name:
    excel file with extension .xlsx
    file to read and write
    column b contains misspelled word.
    column b isn't sorted, we don't want to sort this file.

    In spellings iterate over every row
    Look for match between spellings column b, look in misspelled column b for match (one at most).
    If find match, copy spellings column a to misspelled column a.
    If find match, copy spellings column d to misspelled column d.
    """
    spellings_workbook = load_workbook(spellings_file_name)
    spellings_sheet = spellings_workbook.active

    misspelled_workbook = load_workbook(misspelled_file_name)
    misspelled_sheet = misspelled_workbook.active

    # http://stackoverflow.com/questions/37440855/how-do-i-iterate-through-cells-in-a-specific-column-using-openpyxl-1-6
    status_column = 1

    spellings_sheet_first_non_header_row = 2
    spellings_sheet_misspelled_column = 2
    spellings_sheet_correct_column = 4

    misspelled_sheet_first_non_header_row = 2
    misspelled_sheet_misspelled_column = 2
    misspelled_sheet_correct_column = 4

    # spellings_sheet
    for row in range(spellings_sheet_first_non_header_row, spellings_sheet.max_row + 1):
        spellings_sheet_misspelled_word = spellings_sheet.cell(row=row, column=spellings_sheet_misspelled_column).value
        spelled_word = spellings_sheet.cell(row=row, column=spellings_sheet_correct_column).value
        spelled_status = spellings_sheet.cell(row=row, column=status_column).value

        if ((spellings_sheet_misspelled_word is not None)
            # and (spelled_word is not None)
            # and (spelled_status is not None)):
            and (spelled_word is not None)):

            write_spelling_to_misspelled(misspelled_sheet,
                                         misspelled_sheet_correct_column,
                                         misspelled_sheet_first_non_header_row,
                                         misspelled_sheet_misspelled_column,
                                         row,
                                         spelled_status,
                                         spelled_word,
                                         spellings_sheet_misspelled_word,
                                         status_column)

    misspelled_workbook.save(filename=misspelled_file_name)


def write_spelling_to_misspelled(misspelled_sheet, misspelled_sheet_correct_column,
                                 misspelled_sheet_first_non_header_row, misspelled_sheet_misspelled_column, row,
                                 spelled_status, spelled_word, spellings_sheet_misspelled_word, status_column):

    for misspelled_row in range(misspelled_sheet_first_non_header_row, misspelled_sheet.max_row + 1):
        misspelled_sheet_misspelled_word = misspelled_sheet.cell(row=misspelled_row,
                                                                 column=misspelled_sheet_misspelled_column).value
        misspelled_sheet_correct_column_is_empty = (
        (misspelled_sheet.cell(row=misspelled_row, column=misspelled_sheet_correct_column).value is None)
        or (misspelled_sheet.cell(row=misspelled_row, column=misspelled_sheet_correct_column).value == ''))

        if ((misspelled_sheet_misspelled_word is not None)
            and (misspelled_sheet_misspelled_word == spellings_sheet_misspelled_word)
            and misspelled_sheet_correct_column_is_empty):
            print('spellings_row: ' + str(row) + ', misspelled_row: ' + str(misspelled_row))
            print(spellings_sheet_misspelled_word + ', ' + spelled_word)
            misspelled_sheet.cell(row=misspelled_row, column=misspelled_sheet_correct_column).value = spelled_word
            misspelled_sheet.cell(row=misspelled_row, column=status_column).value = spelled_status


def write_name_matches_to_file(names_in, names_out):
    """
    Reads from 2 excel files. Writes to names_out

    :param names_in:
    excel file with extension .xlsx
    file to read from
    names_in number of rows is <= names_out number of rows
    column 'a' (i.e. 1) contains names that may be spelled correctly or not.
    rows may be sorted alphabetically ascending

    :param names_out:
    excel file with extension .xlsx
    file to read and write
    column 'b' (i.e. 2) contains name.
    rows aren't sorted, we don't want to sort this file.

    In names_in iterate over every row.
    Look for match between names_in column 1, look in names_out column 2 for match (one at most).
    If find match, in names_out column 1 write string "n".
    """
    names_in_workbook = load_workbook(names_in)
    names_in_sheet = names_in_workbook.active

    names_out_workbook = load_workbook(names_out)
    names_out_sheet = names_out_workbook.active

    # http://stackoverflow.com/questions/37440855/how-do-i-iterate-through-cells-in-a-specific-column-using-openpyxl-1-6
    names_in_first_non_header_row = 2
    names_in_column = 1

    names_out_first_non_header_row = 2
    names_out_name_column = 2
    names_out_n_column = 1

    # names_in_sheet
    for row in range(names_in_first_non_header_row, names_in_sheet.max_row + 1):
        name_in = names_in_sheet.cell(row=row, column=names_in_column).value

        if name_in is not None:

            write_name_symbol_to_names_out(name_in,
                                           names_out_sheet,
                                           names_out_first_non_header_row,
                                           names_out_name_column,
                                           names_out_n_column)

    names_out_workbook.save(filename=names_out)


def write_name_symbol_to_names_out(name_in,
                                   names_out_sheet,
                                   names_out_first_non_header_row,
                                   names_out_name_column,
                                   names_out_n_column):

    name_symbol = 'n'

    for row in range(names_out_first_non_header_row, names_out_sheet.max_row + 1):
        name_out = names_out_sheet.cell(row=row, column=names_out_name_column).value
        name_out_is_empty = ((name_out.value is None) or (name_out.value == ''))

        if ((not name_out_is_empty)
            and (name_out == name_in)):
            print('names_out_row: ' + str(row))
            names_out_sheet.cell(row=row, column=names_out_n_column).value = name_symbol
